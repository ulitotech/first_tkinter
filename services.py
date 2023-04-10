import openpyxl as opxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from xls2xlsx import XLS2XLSX
import os
from pathlib import Path
import pandas as pd
from typing import Literal
from tkinter import messagebox as mbox

len_zisu =225

def get_file_path(root:str)->dict:
    """Получение словаря с расположениями рабочих файлов
    root:str - рабочая директория"""
    files_pathes = {}
    for file in os.listdir(root):
        if "ZISU" in file:
            files_pathes["ZISU"] = Path(root, file)
        elif "Список" in file:
            files_pathes["CODES"] = Path(root, file)
        elif "основное" in file and ".xlsx" in file:
            files_pathes["FOR_WORK"] = Path(root, file)
        elif "основное" in file and ".xls" in file:
            files_pathes["MAIN"] = Path(root, file)
    return files_pathes
# get_file_path(WORKING_FOLDER)

def converting_xls_to_xlsx(path:str):
    """Конвертация xls в xlsx
    parh:str - путь к файлу кsонвертации"""
    x2x = XLS2XLSX(path)
    x2x.to_xlsx(path.replace(".xls", ".xlsx"))
# converting_xls_to_xlsx(str(get_file_path(WORKING_FOLDER)["MAIN"]))

def max_row_in_main_df(df)->int:
    """Получение количества строк во фрэйме"""
    i = 0
    for _, row in df.iterrows():
        if type(row[0]) == str:
            break
        else:
            i += 1   
    return i-1
# max_row_in_main_df(df_main)

def max_row_in_main_xl(sheet)->int:
    """Получение количества строк в excel"""
    i = 0
    for row in sheet.iter_rows(min_row = 14, max_row = sheet.max_row, min_col = 1, max_col=1):
        if row[0].value:
            i += 1
        else:
            break
    return i
# max_row_in_main_xl(main_sheet)

def getting_data(path:dict)->list:
    """Формирование списка требуемых данных из сопутствующих таблиц"""

    df_zisu = pd.read_excel(path["ZISU"], usecols=[1, 339, 360, 17],header=None, skiprows=6)
    df_codes = pd.read_excel(path["CODES"], header=None)
    df_main = pd.read_excel(path["FOR_WORK"], usecols=[0,12],header=None, skiprows=13)  
    df_codes.rename({0: 'code', 1: 'name'}, axis=1, inplace=True)
    max_row = max_row_in_main_df(df_main) 
    df_main = df_main[:max_row]
    df_main["serials"] = df_main[12].apply(lambda x: "_"+str(x) if str(x).startswith("0") else x)
    df_main.drop(columns=12, axis=0, inplace=True)
    df_main.set_index(0, inplace=True)
    df_zisu.rename({1:'code', 17: 'serials'}, axis=1, inplace=True)
    result = pd.merge(df_main, df_zisu, on = "serials", how = "left")
    result = pd.merge(result, df_codes, on = "code", how = "left")
    result = result[["serials",339,360,"code","name"]]
    result.drop_duplicates(subset=['serials'],inplace=True)
    result.fillna("нет данных", inplace=True)
    return result.values.tolist()
# getting_data(get_file_path(WORKING_FOLDER))

def formating_cell(cell, style:Literal["header"]|Literal["usual"]|Literal["main"]):
    """Форматирование ячейки"""
    if style=="header":
        font = Font(size=15, name="Times New Roman", bold=True)
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
    elif style == "usual":
        font = Font(size=12, name="Times New Roman", bold=False)
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
    elif style == "main":
        font = Font(size=12, name="Times New Roman", bold=False)
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.font = font   
    cell.alignment = alignment  
    if style != "main":
        cell.border = border  

def creating_header(path:str):
    with open(path["FOR_WORK"], "rb") as main_xl:
        main_wb =  opxl.load_workbook(main_xl, read_only=False)   
        main_sheet = main_wb.active
        header_temp = {23:"Заводской номер прибора учета",24:"1-Величина максимальной мощности; кВт",
                        25:"2-Величина максимальной мощности; кВт",26:"код РЭСа",27:"Наименование РЭС",
                        28:"Сравнение"}
        abc_cell = ("W", "X", "Y", "Z", "AA", "AB")
        
        for i in range(6):
            main_sheet.column_dimensions[f"{abc_cell[i]}"].width = 18.9
            main_sheet.merge_cells(f"{abc_cell[i]}11:{abc_cell[i]}12")
            main_sheet.cell(row=11, column=23+i).value = header_temp[23+i]
            formating_cell(main_sheet.cell(row=11, column=23+i), style="header")
            formating_cell(main_sheet.cell(row=12, column=23+i), style="header")
            main_sheet.cell(row=13, column=23+i).value=23+i
            formating_cell(main_sheet.cell(row=13, column=23+i), style="header")
        main_wb.save(path["FOR_WORK"])
#  creating_header(get_file_path(WORKING_FOLDER))

def adding_data_to_main(path):
    
    zisu_data = getting_data(path)
    global len_zisu
    len_zisu = len(zisu_data)
    abc = ("W", "X", "Y", "Z", "AA", "AB")
    
    with open(path["FOR_WORK"], "rb") as main_xl:
        main_wb =  opxl.load_workbook(main_xl, read_only=False)   
        main_sheet = main_wb.active
        i = 0
        
        merged_list = [str(cell) for cell in main_sheet.merged_cells]
        val = None
        start = end = -1
        for row in zisu_data:
            for j in range(5):
                main_sheet[f"{abc[j]}{14+i}"] = row[j]
                formating_cell(main_sheet[f"{abc[j]}{14+i}"], style="usual")
                
            if "Ярэнерго" in main_sheet[f"V{14+i}"].value:
                if main_sheet[f"R{14+i}"].coordinate in main_sheet.merged_cells:
                    val = main_sheet[f"R{14+i}"].value
                    for m_l in merged_list:
                        if main_sheet[f"R{14+i}"].coordinate == m_l.split(":")[0]:
                            start = int(m_l.split(":")[0][1:])
                            end = int(m_l.split(":")[1][1:])
                            main_sheet.unmerge_cells(m_l)
                if start <= 14+i <= end:
                    main_sheet[f"R{14+i}"].value = val
                formating_cell(main_sheet[f"R{14+i}"], style="main")  
                    
                if main_sheet[f"{abc[1]}{14+i}"].value == 0:
                    main_sheet[f"{abc[1]}{14+i}"].value = main_sheet[f"{abc[2]}{14+i}"].value
                    
                if main_sheet[f"R{14+i}"].value != main_sheet[f"{abc[1]}{14+i}"].value:
                    main_sheet[f"R{14+i}"].fill = PatternFill("solid", fgColor="9400D3")
                    main_sheet[f"{abc[1]}{14+i}"].fill = PatternFill("solid", fgColor="9400D3")
                    main_sheet[f"{abc[-1]}{14+i}"].value = "Не равно"
                else:
                    main_sheet[f"{abc[-1]}{14+i}"].value = "Равно"
    
            formating_cell(main_sheet[f"{abc[-1]}{14+i}"], style="usual")
            i += 1
        main_wb.save(path["FOR_WORK"])      
# adding_data_to_main(get_file_path(WORKING_FOLDER))

def copy_header(workbook:opxl.Workbook, mode:str):
    sheet_to_copy = workbook["Лист1"]
    new_sheet = workbook.copy_worksheet(sheet_to_copy)
    if mode == "no_data":
        new_sheet.title = "no_data"
        workbook["no_data"].delete_rows(14,workbook["no_data"].max_row)
    else:
        new_sheet.title = mode
        workbook[mode].delete_rows(14,workbook[mode].max_row)
# copy_header(main_wb, "no_data")

def no_data(path):
    with open(path["FOR_WORK"], "rb") as main_xl:
        main_wb =  opxl.load_workbook(main_xl, read_only=False)   
        main_sheet = main_wb["Лист1"]
        flag = True
        global len_zisu
        r = 14
        for row in main_sheet.iter_rows(min_row = 13, max_row = 14+len_zisu, min_col = 1, max_col=28):
            if row[-2].value == "нет данных" and "Ярэнерго" in row[-7].value:
                if flag:
        
                    copy_header(main_wb, "no_data")
                    no_data_sheet = main_wb["no_data"]
                    
                    for i in range(len(row)):
                        no_data_sheet.cell(row=r, column=1+i).value = row[i].value
                        if row[i].has_style:
                            no_data_sheet.cell(row=r, column=1+i)._style = row[i]._style
                    flag = False
                    
                else:
                    no_data_sheet = main_wb["no_data"]
                    for i in range(len(row)):
                        no_data_sheet.cell(row=r, column=1+i).value = row[i].value
                        if row[i].has_style:
                            no_data_sheet.cell(row=r, column=1+i)._style = row[i]._style
                r += 1
        main_wb.save(path["FOR_WORK"])        
     main_process(WORKING_FOLDER)            
# no_data(WORKING_FOLDER)

def main_process(path:str, dial, win):
    """Основная рабочая функция, которая запускается после нажатия кнопки""" 
    try:
        converting_xls_to_xlsx(str(get_file_path(path)["MAIN"]))
        creating_header(get_file_path(path))
        adding_data_to_main(get_file_path(path))
        no_data(get_file_path(path))
        different(get_file_path(path))
        mbox.showinfo("INFO", "Документ собран!!!")
        dial.del_param()
    except Exception as err:
        print(err)
# main_process(WORKING_FOLDER)